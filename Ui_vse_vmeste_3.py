# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'c:\Users\Inducyber\Desktop\Python\test\labamain.ui'
#
# Created by: PyQt5 UI code generator 5.15.9
#
# WARNING: Any manual changes made to this file will be lost when pyuic5 is
# run again.  Do not edit this file unless you know what you are doing.

from PyQt5 import QtCore, QtGui, QtWidgets #импорт нескольких модулей Pt
import sys #импорт модуля sys
import openpyxl as op
import qrcode

class Ui_MainWindow(object): 
    def setupUi(self, MainWindow): # определяется пользовательский интерфейс главного окна программы
        MainWindow.setObjectName("MainWindow")
        MainWindow.resize(1200, 500)
        self.centralwidget = QtWidgets.QWidget(MainWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(400, 150, 400, 60))
        font = QtGui.QFont()
        font.setPointSize(26)
        self.label.setFont(font)
        self.label.setAutoFillBackground(False)
        self.label.setObjectName("label")
        self.lineEdit = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit.setGeometry(QtCore.QRect(100, 220, 931, 30))
        self.lineEdit.setObjectName("lineEdit")
        self.GetLibraryCode = QtWidgets.QPushButton(self.centralwidget)
        self.GetLibraryCode.setGeometry(QtCore.QRect(100, 260, 200, 25))
        self.GetLibraryCode.setObjectName("GetLibraryCode")
        self.DigitalLibrary = QtWidgets.QPushButton(self.centralwidget)
        self.DigitalLibrary.setGeometry(QtCore.QRect(300, 260, 200, 25))
        self.DigitalLibrary.setObjectName("DigitalLibrary")
        self.NewArrivals = QtWidgets.QPushButton(self.centralwidget)
        self.NewArrivals.setGeometry(QtCore.QRect(500, 260, 200, 25))
        self.NewArrivals.setObjectName("NewArrivals")
        self.SubResources = QtWidgets.QPushButton(self.centralwidget)
        self.SubResources.setGeometry(QtCore.QRect(700, 260, 200, 25))
        self.SubResources.setObjectName("SubResources")
        self.EducMaterials = QtWidgets.QPushButton(self.centralwidget)
        self.EducMaterials.setGeometry(QtCore.QRect(900, 260, 200, 25))
        self.EducMaterials.setObjectName("EducMaterials")
        self.PersonalArea = QtWidgets.QPushButton(self.centralwidget)
        self.PersonalArea.setGeometry(QtCore.QRect(1000, 30, 100, 25))
        self.PersonalArea.setObjectName("PersonalArea")
        self.Search = QtWidgets.QPushButton(self.centralwidget)
        self.Search.setGeometry(QtCore.QRect(1030, 220, 71, 31))
        self.Search.setObjectName("Search")
        self.bookListWidget = QtWidgets.QListWidget(self.centralwidget)
        self.bookListWidget.setGeometry(QtCore.QRect(100, 300, 931, 150))
        self.bookListWidget.setObjectName("bookListWidget")
        MainWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(MainWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 1200, 21))
        self.menubar.setObjectName("menubar")
        MainWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(MainWindow)
        self.statusbar.setObjectName("statusbar")
        MainWindow.setStatusBar(self.statusbar)

        self.retranslateUi(MainWindow)
        QtCore.QMetaObject.connectSlotsByName(MainWindow)

        self.PersonalArea.clicked.connect(self.openPersonalAreaWindow) # Подвязка кнопки "Ваш кабинет" к функции открытия соответствующего окна
        self.GetLibraryCode.clicked.connect(self.openRegistrationWindow) # Подвязка кнопки "Получить библиотечный код" к функции открытия соответствующего окна
        self.NewArrivals.clicked.connect(self.openNewArrivalsWindow) # Подвязка кнопки "Новые поступления" к функции открытия соответствующего окна
        self.SubResources.clicked.connect(self.openSubResourcesWindow) # Подвязка кнопки "Список подписных ресурсов" к функции открытия соответствующего окна
        self.Search.clicked.connect(self.myFunction) # Подвязка кнопки "Поиск" к функции поиска
        self.lineEdit.returnPressed.connect(self.myFunction) # Подвязка нажатия "Enter" к функции поиска


    def retranslateUi(self, MainWindow): # происходит настройка текстовых надписей на элементах интерфейса
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowTitle(_translate("MainWindow", "MainWindow"))
        self.label.setText(_translate("MainWindow", "НАУЧНАЯ БИБЛИОТЕКА"))
        self.GetLibraryCode.setText(_translate("MainWindow", "Получить библиотечный код"))
        self.DigitalLibrary.setText(_translate("MainWindow", "Электронная библиотека"))
        self.NewArrivals.setText(_translate("MainWindow", "Новые поступления"))
        self.SubResources.setText(_translate("MainWindow", "Список подписных ресурсов"))
        self.EducMaterials.setText(_translate("MainWindow", "Обучающие материалы"))
        self.PersonalArea.setText(_translate("MainWindow", "Ваш кабинет"))
        self.Search.setText(_translate("MainWindow", "Поиск"))

    def openPersonalAreaWindow(self):
        self.window = QtWidgets.QMainWindow()
        self.ui = Ui_PersonalArea()
        self.ui.setupUi(self.window)
        self.window.show()
    
    def openRegistrationWindow(self):
        self.window = QtWidgets.QMainWindow()
        self.ui = Ui_RegistrationWindow()
        self.ui.setupUi(self.window)
        self.window.show()

    def openNewArrivalsWindow(self):
        self.window = QtWidgets.QMainWindow()
        self.ui = Ui_NewArrivalsWindow()
        self.ui.setupUi(self.window)
        self.window.show()

    def openSubResourcesWindow(self):
        self.window = QtWidgets.QMainWindow()
        self.ui = Ui_SubResourcesWindow()
        self.ui.setupUi(self.window)
        self.window.show()
    
    def myFunction(self):
        search_text = self.lineEdit.text()
        filename = 'Список книг библиотеки.xlsx'
        wb = op.load_workbook(filename, data_only=True)
        sheet = wb.active
        max_rows = sheet.max_row

        found_books = []
        for i in range(3, max_rows + 1):
            name_aut = str(sheet.cell(row=i, column=3).value)
            book_name = str(sheet.cell(row=i, column=2).value)
            art = str(sheet.cell(row=i, column=4).value)

            if search_text in name_aut or search_text in book_name or search_text in art:
                found_books.append((book_name, name_aut, art))

        self.bookListWidget.clear()
        if found_books:
            for book in found_books:
                item = QtWidgets.QListWidgetItem(f"{book[0]} - {book[1]} - {book[2]}")
                self.bookListWidget.addItem(item)
                item.setData(QtCore.Qt.UserRole, book)  # Сохраняем информацию о книге в пользовательские данные

class MyMainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super(MyMainWindow, self).__init__()
        self.ui = Ui_MainWindow()
        self.ui.setupUi(self)

       # Подключаем обработчик события для отображения QR-кода при выборе книги из списка
        self.ui.bookListWidget.itemClicked.connect(self.displayQRCode)

    def closeEvent(self, e):
        result = QtWidgets.QMessageBox.question(self,"Окно выхода", "Хотите выйти?", QtWidgets.QMessageBox.Yes | QtWidgets.QMessageBox.No, QtWidgets.QMessageBox.No)
        if result == QtWidgets.QMessageBox.Yes:
            e.accept()
        else:
            e.ignore()

    def displayQRCode(self, item):
        book_info = item.data(QtCore.Qt.UserRole)
        if book_info:
            book_name, _, _ = book_info
            qr = qrcode.QRCode(version=1, error_correction=qrcode.constants.ERROR_CORRECT_L, box_size=10, border=4)
            qr.add_data(book_name)
            qr.make(fit=True)
            img = qr.make_image(fill_color="black", back_color="white")
            img.save("temp_qr_code.png")  # Сохраняем QR-код во временный файл
            self.showQRCode()  # Показываем QR-код

    def showQRCode(self):
        qr_window = QtWidgets.QDialog()
        qr_window.setWindowTitle("QR Code")
        layout = QtWidgets.QVBoxLayout()
        label = QtWidgets.QLabel()
        pixmap = QtGui.QPixmap("temp_qr_code.png")
        label.setPixmap(pixmap)
        layout.addWidget(label)
        qr_window.setLayout(layout)
        qr_window.exec_()

class Ui_PersonalArea(object):
    def setupUi(self, PersonalArea): # определяется пользовательский интерфейс личного окна программы
        PersonalArea.setObjectName("PersonalArea")
        PersonalArea.resize(450, 220)
        self.centralwidget = QtWidgets.QWidget(PersonalArea)
        self.centralwidget.setObjectName("centralwidget")
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(20, 10, 400, 25))
        self.label.setObjectName("label")
        self.lineEdit = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit.setGeometry(QtCore.QRect(20, 40, 400, 25))
        self.lineEdit.setObjectName("lineEdit")
        self.label_2 = QtWidgets.QLabel(self.centralwidget)
        self.label_2.setGeometry(QtCore.QRect(20, 80, 400, 25))
        self.label_2.setObjectName("label_2")
        self.lineEdit_2 = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit_2.setGeometry(QtCore.QRect(20, 110, 400, 25))
        self.lineEdit_2.setObjectName("lineEdit_2")
        self.pushButton = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton.setGeometry(QtCore.QRect(20, 150, 120, 30))
        self.pushButton.setObjectName("pushButton")
        self.pushButton_2 = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton_2.setGeometry(QtCore.QRect(150, 150, 270, 30))
        self.pushButton_2.setObjectName("pushButton_2")
        PersonalArea.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(PersonalArea)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 450, 21))
        self.menubar.setObjectName("menubar")
        PersonalArea.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(PersonalArea)
        self.statusbar.setObjectName("statusbar")
        PersonalArea.setStatusBar(self.statusbar)

        self.retranslateUi(PersonalArea)
        QtCore.QMetaObject.connectSlotsByName(PersonalArea)

    def retranslateUi(self, PersonalArea):
        _translate = QtCore.QCoreApplication.translate
        PersonalArea.setWindowTitle(_translate("PersonalArea", "MainWindow"))
        self.label.setText(_translate("PersonalArea", "Библиотечный код или email"))
        self.label_2.setText(_translate("PersonalArea", "Пароль (дата рождения в формате ГГГГММДД):"))
        self.pushButton.setText(_translate("PersonalArea", "Войти"))
        self.pushButton_2.setText(_translate("PersonalArea", "Забыли свой библиотечный код?"))

class Ui_RegistrationWindow(object): 
    def setupUi(self, RegistrationWindow): # определяется пользовательский интерфейс регистрационного окна программы
        RegistrationWindow.setObjectName("RegistrationWindow")
        RegistrationWindow.resize(450, 550)
        self.centralwidget = QtWidgets.QWidget(RegistrationWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(65, 30, 310, 40))
        font = QtGui.QFont()
        font.setPointSize(15)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.label_2 = QtWidgets.QLabel(self.centralwidget)
        self.label_2.setGeometry(QtCore.QRect(30, 80, 300, 25))
        font = QtGui.QFont()
        font.setPointSize(8)
        self.label_2.setFont(font)
        self.label_2.setObjectName("label_2")
        self.lineEdit = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit.setGeometry(QtCore.QRect(30, 110, 400, 25))
        self.lineEdit.setObjectName("lineEdit")
        self.label_3 = QtWidgets.QLabel(self.centralwidget)
        self.label_3.setGeometry(QtCore.QRect(30, 140, 300, 25))
        font = QtGui.QFont()
        font.setPointSize(8)
        self.label_3.setFont(font)
        self.label_3.setObjectName("label_3")
        self.lineEdit_2 = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit_2.setGeometry(QtCore.QRect(30, 170, 400, 25))
        self.lineEdit_2.setObjectName("lineEdit_2")
        self.label_4 = QtWidgets.QLabel(self.centralwidget)
        self.label_4.setGeometry(QtCore.QRect(30, 200, 300, 25))
        font = QtGui.QFont()
        font.setPointSize(8)
        self.label_4.setFont(font)
        self.label_4.setObjectName("label_4")
        self.lineEdit_3 = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit_3.setGeometry(QtCore.QRect(30, 230, 400, 25))
        self.lineEdit_3.setObjectName("lineEdit_3")
        self.label_5 = QtWidgets.QLabel(self.centralwidget)
        self.label_5.setGeometry(QtCore.QRect(30, 260, 300, 25))
        font = QtGui.QFont()
        font.setPointSize(8)
        self.label_5.setFont(font)
        self.label_5.setObjectName("label_5")
        self.lineEdit_4 = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit_4.setGeometry(QtCore.QRect(30, 290, 400, 25))
        self.lineEdit_4.setObjectName("lineEdit_4")
        self.label_6 = QtWidgets.QLabel(self.centralwidget)
        self.label_6.setGeometry(QtCore.QRect(30, 320, 300, 25))
        font = QtGui.QFont()
        font.setPointSize(8)
        self.label_6.setFont(font)
        self.label_6.setObjectName("label_6")
        self.lineEdit_5 = QtWidgets.QLineEdit(self.centralwidget)
        self.lineEdit_5.setGeometry(QtCore.QRect(30, 350, 400, 25))
        self.lineEdit_5.setObjectName("lineEdit_5")
        self.label_7 = QtWidgets.QLabel(self.centralwidget)
        self.label_7.setGeometry(QtCore.QRect(30, 380, 300, 25))
        font = QtGui.QFont()
        font.setPointSize(8)
        self.label_7.setFont(font)
        self.label_7.setObjectName("label_7")
        self.comboBox = QtWidgets.QComboBox(self.centralwidget)
        self.comboBox.setGeometry(QtCore.QRect(30, 410, 400, 25))
        self.comboBox.setObjectName("comboBox")
        self.pushButton = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton.setGeometry(QtCore.QRect(125, 460, 200, 41))
        font = QtGui.QFont()
        font.setPointSize(8)
        self.pushButton.setFont(font)
        self.pushButton.setObjectName("pushButton")
        RegistrationWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(RegistrationWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 450, 21))
        self.menubar.setObjectName("menubar")
        RegistrationWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(RegistrationWindow)
        self.statusbar.setObjectName("statusbar")
        RegistrationWindow.setStatusBar(self.statusbar)

        self.comboBox.setCurrentIndex(0)
        self.comboBox.addItems('студенты магистранты аспиранты докторанты преподаватели'.split())

        self.retranslateUi(RegistrationWindow)
        QtCore.QMetaObject.connectSlotsByName(RegistrationWindow)

    def retranslateUi(self, RegistrationWindow):
        _translate = QtCore.QCoreApplication.translate
        RegistrationWindow.setWindowTitle(_translate("RegistrationWindow", "MainWindow"))
        self.label.setText(_translate("RegistrationWindow", "Регистрация библиотечного кода"))
        self.label_2.setText(_translate("RegistrationWindow", "Фамилия:"))
        self.label_3.setText(_translate("RegistrationWindow", "Имя:"))
        self.label_4.setText(_translate("RegistrationWindow", "Отчество (при наличии):"))
        self.label_5.setText(_translate("RegistrationWindow", "Дата рождения:"))
        self.label_6.setText(_translate("RegistrationWindow", "email:"))
        self.label_7.setText(_translate("RegistrationWindow", "Выберите категорию:"))
        self.pushButton.setText(_translate("RegistrationWindow", "Зарегистрироваться"))

class Ui_NewArrivalsWindow(object):
    def setupUi(self, NewArrivalsWindow): # определяется пользовательский интерфейс новых поступлений окна программы
        NewArrivalsWindow.setObjectName("NewArrivalsWindow")
        NewArrivalsWindow.resize(400, 220)
        self.centralwidget = QtWidgets.QWidget(NewArrivalsWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(107, 30, 185, 40))
        font = QtGui.QFont()
        font.setPointSize(15)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.label_2 = QtWidgets.QLabel(self.centralwidget)
        self.label_2.setGeometry(QtCore.QRect(20, 70, 130, 20))
        self.label_2.setObjectName("label_2")
        self.comboBox = QtWidgets.QComboBox(self.centralwidget)
        self.comboBox.setGeometry(QtCore.QRect(20, 90, 360, 25))
        self.comboBox.setObjectName("comboBox")
        self.pushButton = QtWidgets.QPushButton(self.centralwidget)
        self.pushButton.setGeometry(QtCore.QRect(20, 130, 120, 30))
        self.pushButton.setObjectName("pushButton")
        NewArrivalsWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(NewArrivalsWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 400, 21))
        self.menubar.setObjectName("menubar")
        NewArrivalsWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(NewArrivalsWindow)
        self.statusbar.setObjectName("statusbar")
        NewArrivalsWindow.setStatusBar(self.statusbar)

        self.comboBox.setCurrentIndex(0)
        self.comboBox.addItems(['за этот месяц','с прошлого месяца','с позапрошлого месяца'])

        self.retranslateUi(NewArrivalsWindow)
        QtCore.QMetaObject.connectSlotsByName(NewArrivalsWindow)

    def retranslateUi(self, NewArrivalsWindow):
        _translate = QtCore.QCoreApplication.translate
        NewArrivalsWindow.setWindowTitle(_translate("NewArrivalsWindow", "MainWindow"))
        self.label.setText(_translate("NewArrivalsWindow", "Новые поступления"))
        self.label_2.setText(_translate("NewArrivalsWindow", "Новые книги в каталоге"))
        self.pushButton.setText(_translate("NewArrivalsWindow", "Показать"))

class Ui_SubResourcesWindow(object):
    def setupUi(self, SubResourcesWindow): # определяется пользовательский интерфейс подписных ресурсов окна программы
        SubResourcesWindow.setObjectName("SubResourcesWindow")
        SubResourcesWindow.resize(500, 500)
        self.centralwidget = QtWidgets.QWidget(SubResourcesWindow)
        self.centralwidget.setObjectName("centralwidget")
        self.label = QtWidgets.QLabel(self.centralwidget)
        self.label.setGeometry(QtCore.QRect(120, 40, 261, 25))
        font = QtGui.QFont()
        font.setPointSize(15)
        self.label.setFont(font)
        self.label.setObjectName("label")
        self.comboBox = QtWidgets.QComboBox(self.centralwidget)
        self.comboBox.setGeometry(QtCore.QRect(10, 80, 235, 25))
        self.comboBox.setObjectName("comboBox")
        self.comboBox_2 = QtWidgets.QComboBox(self.centralwidget)
        self.comboBox_2.setGeometry(QtCore.QRect(255, 80, 235, 25))
        self.comboBox_2.setObjectName("comboBox_2")
        self.plainTextEdit = QtWidgets.QPlainTextEdit(self.centralwidget)
        self.plainTextEdit.setGeometry(QtCore.QRect(10, 110, 480, 330))
        self.plainTextEdit.setObjectName("plainTextEdit")
        SubResourcesWindow.setCentralWidget(self.centralwidget)
        self.menubar = QtWidgets.QMenuBar(SubResourcesWindow)
        self.menubar.setGeometry(QtCore.QRect(0, 0, 500, 21))
        self.menubar.setObjectName("menubar")
        SubResourcesWindow.setMenuBar(self.menubar)
        self.statusbar = QtWidgets.QStatusBar(SubResourcesWindow)
        self.statusbar.setObjectName("statusbar")
        SubResourcesWindow.setStatusBar(self.statusbar)
        
        self.comboBox.setCurrentIndex(0)
        self.comboBox.addItems(['все темы','авиационная и космическая техника','автоматика и управление','архитектура и строительство',
        'гуманитарные науки','культура и искусство','социальные науки','физико-математические науки'])
        self.comboBox_2.setCurrentIndex(0)
        self.comboBox_2.addItems(['все материалы','база знаний','видео','книги','лекции','материалы концеренций','технические документы','справочные издания'])

        self.retranslateUi(SubResourcesWindow)
        QtCore.QMetaObject.connectSlotsByName(SubResourcesWindow)

    def retranslateUi(self, SubResourcesWindow):
        _translate = QtCore.QCoreApplication.translate
        SubResourcesWindow.setWindowTitle(_translate("SubResourcesWindow", "MainWindow"))
        self.label.setText(_translate("SubResourcesWindow", "Список подписных ресурсов"))

if __name__ == "__main__":
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = MyMainWindow()
    MainWindow.show()
    sys.exit(app.exec_())